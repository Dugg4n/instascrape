from flask import Flask, render_template, request, send_file
import requests, json, re, time, io
import browser_cookie3
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

app = Flask(__name__)

POST_QUERY_HASH = "97b41c52301f77ce508f55e66d17620e"
COMMENTS_PER_PAGE = 100

def extract_shortcode(url):
    m = re.search(r"instagram\.com/(?:reel|p)/([^/?]+)", url)
    return m.group(1) if m else None

def build_headers(shortcode, cookies_str):
    return {
        "User-Agent": "Mozilla/5.0",
        "Accept": "/",
        "Accept-Language": "en-US,en;q=0.9",
        "X-Requested-With": "XMLHttpRequest",
        "X-IG-App-ID": "936619743392459",
        "Referer": f"https://www.instagram.com/p/{shortcode}/",
        "Cookie": cookies_str
    }

def graphql_request(query_hash, variables, headers):
    var_str = json.dumps(variables, separators=(",", ":"))
    url = f"https://www.instagram.com/graphql/query/?query_hash={query_hash}&variables={requests.utils.quote(var_str)}"
    r = requests.get(url, headers=headers)
    r.raise_for_status()
    return r.json()

def fetch_comments(shortcode, headers):
    all_comments = []
    has_next = True
    cursor = ""

    while has_next:
        vars = {"shortcode": shortcode, "first": COMMENTS_PER_PAGE}
        if cursor:
            vars["after"] = cursor
        data = graphql_request(POST_QUERY_HASH, vars, headers)

        edge_info = data["data"]["shortcode_media"]["edge_media_to_parent_comment"]
        edges = edge_info["edges"]

        for edge in edges:
            node = edge["node"]
            user = node["owner"]["username"]
            text = node["text"]
            all_comments.append({"user": user, "comment": text, "reply_to": None})

            if node.get("edge_threaded_comments"):
                for reply in node["edge_threaded_comments"]["edges"]:
                    r_node = reply["node"]
                    r_user = r_node["owner"]["username"]
                    r_text = r_node["text"]
                    all_comments.append({"user": r_user, "comment": r_text, "reply_to": user})

        page_info = edge_info["page_info"]
        has_next = page_info["has_next_page"]
        cursor = page_info["end_cursor"]
        time.sleep(0.5)

    return all_comments

def get_cookies_auto():
    browsers = ['chrome', 'chromium', 'brave', 'firefox']
    cookies_dict = {}
    for browser in browsers:
        try:
            cj = getattr(browser_cookie3, browser)(domain_name='instagram.com')
            for c in cj:
                cookies_dict[c.name] = c.value
            if cookies_dict:
                return cookies_dict
        except Exception:
            continue
    return None

def export_comments(comments, format_type, filename_base):
    filename_base = filename_base.strip() or "comments"

    if format_type == "txt":
        output = io.BytesIO()
        content = ""
        for c in comments:
            prefix = "  ↳ " if c["reply_to"] else ""
            content += f"{prefix}[{c['user']}] {c['comment']}\n"
        output.write(content.encode("utf-8"))
        output.seek(0)
        return output, f"{filename_base}.txt"

    elif format_type == "json":
        output = io.BytesIO()
        json_data = json.dumps(comments, indent=2, ensure_ascii=False)
        output.write(json_data.encode("utf-8"))
        output.seek(0)
        return output, f"{filename_base}.json"

    elif format_type == "xlsx":
        output = io.BytesIO()
        wb = Workbook()
        ws = wb.active
        ws.title = "Instagram Comments"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        align = Alignment(vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        headers = ["Username", "Comment", "ReplyTo"]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align
            cell.border = thin_border

        for c in comments:
            row = [c["user"], c["comment"], c["reply_to"] or ""]
            ws.append(row)

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=3):
            for cell in row:
                cell.alignment = align
                cell.border = thin_border

        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 60)

        wb.save(output)
        output.seek(0)
        return output, f"{filename_base}.xlsx"

    else:
        return None, ""

@app.route("/", methods=["GET", "POST"])
def index():
    message = ""
    comments = []
    if request.method == "POST":
        url = request.form["url"]
        format_type = request.form["format"]
        filename_base = request.form["filename"]
        shortcode = extract_shortcode(url)
        if not shortcode:
            message = "Invalid Instagram URL."
        else:
            cookies = get_cookies_auto()
            if not cookies:
                message = "Failed to detect browser cookies."
            else:
                needed = ['sessionid', 'ds_user_id', 'mid', 'csrftoken']
                cookie_str = "; ".join(f"{k}={cookies.get(k, '')}" for k in needed)
                headers = build_headers(shortcode, cookie_str)
                try:
                    comments = fetch_comments(shortcode, headers)

                    if "download" in request.form:
                        export_data, filename = export_comments(comments, format_type, filename_base)
                        mimetypes = {
                            "txt": "text/plain",
                            "json": "application/json",
                            "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        }
                        return send_file(export_data, as_attachment=True, download_name=filename, mimetype=mimetypes[format_type])

                except Exception as e:
                    message = f"Error: {str(e)}"
    return render_template("index.html", message=message, comments=comments)

if __name__ == "__main__":
    app.run(debug=True)
